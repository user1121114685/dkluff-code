#!/root/anaconda2/bin/python2
# -*- coding: UTF-8 -*-

"""
problem version 1: 母件编码 和 子件明细 STACK
update： 2018/01/12
"""
import sys


import glob
import os
import re
import time


import pandas as pd
import numpy as np

import openpyxl as pyxls
from openpyxl.utils.dataframe import dataframe_to_rows
import ConfigParser

import pdb
import pickle as pk
#import codecs
#import csv
#import xlrd

"""
http://openpyxl.readthedocs.io/en/latest/tutorial.html#accessing-one-cell

http://pandas-docs.github.io/pandas-docs-travis/dsintro.html#dataframe

https://www.joinquant.com/post/1980?f=study&m=python

https://pyinstaller.readthedocs.io/en/stable/usage.html

"""
#[base]

#HEAD_SUBTAB = "母件编号,级别,子件行号,工序行号,工序名称,子件编码,子件名称,子件规格,子件计量单位,基本用量,基础数量,子件损耗率,辅助单位,换算率,辅助基本用量,辅助使用量,固定用量,供应类型,使用数量,子件生效日,子件失效日,偏置期,计划比例,产出品,产出类型,成本相关,可选否,选择规则,仓库名称,领料部门名称,定位符,备注"

#HEAD_MTAB = "母件编码,母件名称,母件规格,母件计量单位,母件损耗率,版本代号,版本说明,版本日期,状态"

#MC_STR = "母件编码"
CFG_FILENAME = './base.cfg'
XLSFILES='./xlsfiles/'
RESULTS='./results/'



def todf(atab,cols):
    cl = len(cols)
    ntab = []
    for i in atab:
        if len(i)<cl:
            i = i+[ "" for x in range(cl-len(i))]
        ntab.append(i[:cl])
    df = pd.DataFrame(ntab,columns=cols)

    return df


def reader_pyxls(xlsfile):
    #read xls and format#1
    wb = pyxls.load_workbook(filename = xlsfile)
    ws = wb.active

    ntab = []
    
    #read xls and encode
    for r in ws.iter_rows():
        ntab.append(list(x.value.encode('utf-8') if type(x.value) == type(u'') else x.value
                         for x in r ))

    return ntab


def save_df_xls(dflist, xls):
    #dflist = [(df,sheetname),...]

    wb = pyxls.Workbook()

    for i in range(len(dflist)):

        ws = wb.create_sheet(dflist[i][1])

        for r in dataframe_to_rows(dflist[i][0], index=True, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

    wb.save(xls)

def table_class(dt,kw=[]):

    dt_dic={}
    for w in kw:
        dt_dic[w]=[]

    

    for t in dt:
        rowshift = 0
        for row in t:   
            try:
                k = row[0]
                if k in kw:
                    dt_dic[k].append(t[rowshift:])
                    break
            except:
                continue
            rowshift+=1
    
    return dt_dic

def proc_tablehead(tb,keyword):
    for row in tb:
        try:
            if keyword == row[0]:
                return row
        except:
            pass
    return

    

def proc_mtable(dtdic,keyword="",keyword2=""):
    """
        mtable - stack table:
        keyword h2 h3 h4
        1        1 1  s

        keyword2 hh2 hh3 hh4
        1        2  3     5
    """
    
    if not dtdic.has_key(keyword):
        print "No such table - keyword error!"
        return
    

    subtable = []
    subheadcols = []
    mitable = []

    for tb in dtdic[keyword]:
        mc = ""
        for i in range(len(tb)):
            #row - tb[i]
            if len(tb[i]) <= 0: continue
            try:
                if mc != "" and ("+" in tb[i][0]):
                    subtable.append([mc]+tb[i])

                if keyword in tb[i][0]:
                    m_i = tb[i+1]
                    mitable.append(m_i)
                    mc = m_i[0]
            except:
                continue

    if len(subtable)>0:
        subheadcols = proc_tablehead(dtdic[keyword][0],keyword2)

    return mitable,subtable,[keyword]+subheadcols

def proc_pricetable(dtdic,keyword=""):

    if not dtdic.has_key(keyword):
        print "No such table - keyword error!"
        return

    pricetable = []
    headcols = []
    #pdb.set_trace()

    for tb in dtdic[keyword]:
        for r in tb[1:]:
            try:
                if len(r)>0 and (r[0] != None):
                    pricetable.append(r)
            except:
                continue

    if len(pricetable)>0:
        headcols = proc_tablehead(dtdic[keyword][0],keyword)

    return pricetable,headcols

def proc_rdrop(df,cols):
    for k in df.columns:
        if k not in cols:
            try:
                df=df.drop(k,axis=1)
            except:
                pass
            
    return df

def cleancfg(fname):
    content = open(fname).read()  
    #Window下用记事本打开配置文件并修改保存后，编码为UNICODE或UTF-8的文件的文件头  
    #会被相应的加上\xff\xfe（\xff\xfe）或\xef\xbb\xbf，然后再传递给ConfigParser解析的时候会出错  
    #，因此解析之前，先替换掉  
    content = re.sub(r"\xfe\xff","", content)  
    content = re.sub(r"\xff\xfe","", content)  
    content = re.sub(r"\xef\xbb\xbf","", content)  
    open(fname, 'w').write(content)


def printwin(s):
    try:
        print s.decode("UTF-8").encode("GB18030")
    except:
        print s


def task_a():
    if not os.path.exists(RESULTS):
        os.makedirs(RESULTS)
    
    if not os.path.exists(XLSFILES):
        os.makedirs(XLSFILES)

    if len(glob.glob(XLSFILES+'*.xls*')) == 0:
        printwin("把原始文件放到以下目录后重新运行：")
        printwin(os.getcwd()+XLSFILES)
        return

    #read config
    cleancfg(CFG_FILENAME)
    config = ConfigParser.RawConfigParser()
    config.read([CFG_FILENAME])

    MTABLE_KEYWORD = config.get('base','MTABLE_KEYWORD')
    SUBTABLE_KEYWORD = config.get('base','SUBTABLE_KEYWORD')
    PRICIE_KEYWORD = config.get('base','PRICIE_KEYWORD')
    FINAL_COLS = config.get('base','FINAL_COLS').split()
    

    datatables = []
    printwin("开始读取文件，可能需要几分钟...")

    
    for f in glob.glob(XLSFILES+'*.xls*'):
        printwin("正在读取： ")
        printwin(f)
        datatables.append(reader_pyxls(f))
    

    #pk.dump([datatables],open("dt.pk","wb"))

    printwin( "加工中。。。")
    tabdic = table_class(datatables,[PRICIE_KEYWORD,MTABLE_KEYWORD])
    pt,hcols = proc_pricetable(tabdic,keyword=PRICIE_KEYWORD)
    
    mit,subt,subthcols = proc_mtable(tabdic,keyword=MTABLE_KEYWORD,keyword2=SUBTABLE_KEYWORD)

    ptf=todf(pt,hcols)
    ptf_indexed = ptf.set_index('存货编码')

    subtf=todf(subt,subthcols)

    #mit 母件信息 无表头
    mitf = pd.DataFrame(mit)

    bomtmp = subtf.join(ptf_indexed,on='子件编码')
    bom = proc_rdrop(bomtmp,FINAL_COLS)
    bom = bom.assign(total = lambda x: x["价格"]*x["使用数量"] )

    bom_sum = bom.groupby("母件编码")["母件编码","total"].sum()
    bom_sub_sum = bom.groupby("子件编码")["子件编码","total"].sum()

    #nb=bom.rename(index=str, columns={"母件编码": "mid", "价格": "price","使用数量":"cout"})
    #pdb.set_trace()
    
    printwin("哈哈哈！生成完毕，保存文件中...")
    xlsname = RESULTS+time.strftime('%y%b%d_%H%M%S')+'.xlsx'
    save_df_xls([(bom,"bom") ,(bom_sum,"bom_sum"),(bom_sub_sum,"bom_sub_sum"),(mitf,"m_ids")],xlsname)


    printwin("任务完成！！！")
    #pdb.set_trace()


#-------------------------------
def testdump():
    datatables = pk.load(open("dt.pk"))[0]

    cleancfg(CFG_FILENAME)
    config = ConfigParser.RawConfigParser()
    config.read([CFG_FILENAME])
    #HEAD_SUBTAB = config.get('base', 'HEAD_SUBTAB')
    #HEAD_MTAB = config.get('base','HEAD_MTAB')
    MTABLE_KEYWORD = config.get('base','MTABLE_KEYWORD')
    SUBTABLE_KEYWORD = config.get('base','SUBTABLE_KEYWORD')
    PRICIE_KEYWORD = config.get('base','PRICIE_KEYWORD')

    dt=datatables
    dtdic = table_class(dt)
    
    pdb.set_trace()





"""
def test():
    t=5

    if t == 2:
        m,s = reader_pyxls("a.xlsx")
        pk.dump([m,s],open("reader_pyxl.pk","wb"))

    if t == 3:
        m,s = pk.load(open("reader_pyxl.pk"))
        sdf = todf(s,HEAD_SUBTAB)
        mdf = todf(m,HEAD_MTAB)
        pk.dump([mdf,sdf],open("dfs.pk","wb"))

    if t == 4:
        mdf,sdf = pk.load(open("dfs.pk"))
        print sdf[:3]
        print sdf[sdf['母件编号'] == "CPEA20001"]
        print sdf.drop(['备注','领料部门名称'],axis=1)[:3]
        #print sdf.drop([u'备注'.encode("utf-8")],axis=1)[:3] #windows 下交互模式的必要方法
        #save_df_xls(sdf,"b.xlsx",sheetname="BOM")

    main()
    
    print os.getcwd()
    print HEAD_MTAB.decode("UTF-8").encode("GB18030")
    print HEAD_SUBTAB.decode("UTF-8").encode("GB18030")
    
    if os.path.exists(CFG_FILENAME):
        print "--------------"
        hs,hm,mc = readconfig()

        print hs.decode("UTF-8").encode("GB18030")
        print hm.decode("UTF-8").encode("GB18030")
"""

def main():
    task = "a"
    if len(sys.argv)>1:
        task = sys.argv[1]
    
    print "run task: ",task
    if task == "a":
        task_a()
    

if __name__ == "__main__":
    #testdump()
    main()






