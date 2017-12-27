#!/root/anaconda2/bin/python2
# -*- coding: UTF-8 -*-

import sys
import csv
import xlrd
import codecs
import glob

import pandas as pd
import numpy as np

import pickle as pk
import openpyxl as pyxls
from openpyxl.utils.dataframe import dataframe_to_rows


"""
http://openpyxl.readthedocs.io/en/latest/tutorial.html#accessing-one-cell

http://pandas-docs.github.io/pandas-docs-travis/dsintro.html#dataframe

https://www.joinquant.com/post/1980?f=study&m=python

"""

HEAD_SUBTAB = "母件编号,级别,子件行号,工序行号,工序名称,子件编码,子件名称,子件规格,子件计量单位,基本用量,基础数量,子件损耗率,辅助单位,换算率,辅助基本用量,辅助使用量,固定用量,供应类型,使用数量,子件生效日,子件失效日,偏置期,计划比例,产出品,产出类型,成本相关,可选否,选择规则,仓库名称,领料部门名称,定位符,备注"

HEAD_MTAB = "母件编码,母件名称,母件规格,母件计量单位,母件损耗率,版本代号,版本说明,版本日期,状态"

MC_STR = "母件编码"

XLSFILES='./xlsfiles/'
RESULTS='./results/'


def main():
    for f in glob.glob(SRCDIR+'*.xls*'):
        print f


def todf(atab,tabhead):
    cols = tabhead.split(",")
    cl = len(cols)
    ntab = []
    for i in atab:
        if len(i)<cl:
            i = i+[ "" for x in range(cl-len(i))]
        ntab.append(i[:cl])
    df = pd.DataFrame(ntab,columns=cols)

    return df

def reader_xlrd(xlsfile):
    #import openpyxl as pyxls
    #wb = pyxls.load_workbook(filename = xlsfile)
    wb = xlrd.open_workbook(xlsfile)
    ws = wb.sheets()[0]

    ntab = []
    subtable = []
    mtable = []

    #read xls and encode
    for rownum in xrange(ws.nrows):
        ntab.append(list(x.encode('utf-8') if type(x) == type(u'') else x
                         for x in ws.row_values(rownum)))
    #format data
    mc = ""
    for i in range(len(ntab)):
        if mc != "" and ("+" in ntab[i][0]):
            subtable.append([mc]+ntab[i])

        if MC_STR in ntab[i][0]:
            m_i = ntab[i+1]
            mtable.append(m_i)
            mc = m_i[0]

    return mtable,subtable


def reader_pyxls(xlsfile):
    wb = pyxls.load_workbook(filename = xlsfile)
    ws = wb.active

    ntab = []
    subtable = []
    mtable = []

    #read xls and encode
    for r in ws.iter_rows():
        ntab.append(list(x.value.encode('utf-8') if type(x.value) == type(u'') else x.value
                         for x in r ))
    #format data
    mc = ""
    for i in range(len(ntab)):
        if mc != "" and ("+" in ntab[i][0]):
            subtable.append([mc]+ntab[i])

        if MC_STR in ntab[i][0]:
            m_i = ntab[i+1]
            mtable.append(m_i)
            mc = m_i[0]

    return mtable,subtable


def save_df_xls(df,xls,sheetname="pysheet01",wb=""):

    if not wb: wb = pyxls.Workbook()
    ws = wb.create_sheet(sheetname)

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    wb.save(xls)







#-------------------------------

def test():
    t=3

    if t == 2:
        m,s = reader_pyxls("a.xlsx")
        pk.dump([m,s],open("reader_pyxl.pk","wb"))

    if t == 3:
        m,s = pk.load(open("reader_pyxl.pk"))
        sdf = todf(s,HEAD_SUBTAB)
        mdf = todf(m,HEAD_MTAB)
        pk.dump([mdf,sdf],open("dfs.pk","wb"))

    if t == 4:
        mdf,sdf = pk.load(open("dfs.pk"))
        print sdf[:3]
        print sdf[sdf['母件编号'] == "CPEA20001"]
        #save_df_xls(sdf,"b.xlsx",sheetname="BOM")


    return

if __name__ == "__main__":

	test()






